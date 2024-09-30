import openpyxl
import requests
#从dnspod导出域名列表并且另存为xlsx
#此脚本用来从域名列表xlsx中导出所有A记录和CNAME记录的域名
def check_subdomains(filepath):
    # 加载工作簿
    workbook = openpyxl.load_workbook(filename=filepath)

    # 遍历所有工作表（Sheet）
    for sheet_name in workbook.sheetnames:
        domain = sheet_name  # 假设sheet的名字就是域名

        # 获取当前工作表
        sheet = workbook[sheet_name]

        # 遍历B列，从第二行开始（假设第一行是标题）
        for row in range(2, sheet.max_row + 1):
            # 获取A列的值
            record_type_cell = sheet.cell(row, 1)  # A列对应索引1

            # 检查A列的值是否为“A”或“CNAME”
            if record_type_cell.value in ["A", "CNAME"]:
                # 获取B列的值
                subdomain_cell = sheet.cell(row, 2)  # B列对应索引2

                # 检查单元格值是否不是"主机记录"或"@"
                if subdomain_cell.value not in ["主机记录", "@"]:
                    subdomain_value = subdomain_cell.value

                    # 如果单元格为空，则跳过
                    if subdomain_value is None:
                        continue

                    subdomain = subdomain_value
                    dns_record = f"{subdomain}.{domain}"
                    print(f"{dns_record}")

if __name__ == "__main__":
    filepath = "dns.xlsx"  # Excel文件路径
    check_subdomains(filepath)
